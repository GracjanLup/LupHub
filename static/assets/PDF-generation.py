import logging
import os
import re
import shutil
import subprocess
import tempfile
import threading
from io import BytesIO
from typing import Callable, Iterable, List, Optional

from PyPDF2 import PdfMerger, PdfReader, PdfWriter

try:
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
except ImportError:  # pragma: no cover - środowisko bez reportlab
    canvas = None
    pdfmetrics = None
    TTFont = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

logger = logging.getLogger(__name__)

BASE_DIR = os.environ.get("WANO_BASE_DIR", "/home/wano")
EXCEL_PATH = os.environ.get("WANO_EXCEL_PATH") or os.path.join(BASE_DIR, "cenniki", "Cennik Partnera WANO.xlsm")
EXPORT_DIR = os.environ.get("WANO_EXPORT_DIR") or os.path.join(BASE_DIR, "ex")
PDFY_DIR = os.environ.get("WANO_PDFY_DIR") or os.path.join(BASE_DIR, "pdfy")
OUTPUT_DIR_PL = os.environ.get("WANO_OUTPUT_PL_DIR") or os.path.join(BASE_DIR, "cennikiPDF-PL")
OUTPUT_DIR_EN = os.environ.get("WANO_OUTPUT_EN_DIR") or os.path.join(BASE_DIR, "cennikiPDF-EN")
OUTPUT_FILE_PL = os.path.join(OUTPUT_DIR_PL, "Cennik B2B WANO.pdf")
OUTPUT_FILE_EN = os.path.join(OUTPUT_DIR_EN, "Price List B2B WANO.pdf")
FOOTER_LEFT_TEXT = "          tel. +48 61 307 22 35"
FOOTER_RIGHT_TEXT = "biuro@wano.pl          "
FOOTER_FONT = "Arial"
FOOTER_FALLBACK_FONT = "Helvetica"
FOOTER_FONT_SIZE = 8
FOOTER_MARGIN_X = 60
FOOTER_MARGIN_Y = 40
# Indeksy stron do pominięcia (0-based)
FOOTER_START_PAGE_INDEX = 4  # zaczynamy od 5. strony
FOOTER_SKIP_LAST = 1
FOOTER_PREFERRED_FONT_PATHS = [
    "/usr/share/fonts/truetype/msttcorefonts/Arial.ttf",
    "/usr/share/fonts/truetype/msttcorefonts/arial.ttf",
    "/usr/share/fonts/truetype/msttcorefonts/ArialMT.ttf",
    "/usr/share/fonts/truetype/msttcorefonts/ArialUnicode.ttf",
]
_FOOTER_FONT_IN_USE: Optional[str] = None


class GenerationError(Exception):
    """Raised when PDF generation cannot be completed."""


class GenerationCancelled(GenerationError):
    """Raised when generation was cancelled by user."""


def _check_cancel(cancel_event: Optional[threading.Event], message: str = "Generowanie przerwane."):
    if cancel_event and cancel_event.is_set():
        raise GenerationCancelled(message)


def _validate_assets(language_token: str, excel_path: Optional[str] = None):
    source_excel = os.path.abspath(excel_path or EXCEL_PATH)
    if not os.path.exists(source_excel):
        raise GenerationError(f"Nie znaleziono pliku Excela: {source_excel}")

    required = [
        os.path.join(PDFY_DIR, f"Start{language_token}.pdf"),
        os.path.join(PDFY_DIR, f"End{language_token}.pdf"),
    ]
    for path in required:
        if not os.path.exists(path):
            raise GenerationError(f"Brak wymaganych plików PDF: {path}")


def _merge_pdf_list(paths: Iterable[str], output_file: str, cancel_event: Optional[threading.Event] = None) -> str:
    _check_cancel(cancel_event)
    os.makedirs(os.path.dirname(os.path.abspath(output_file)) or ".", exist_ok=True)
    merger = PdfMerger()
    try:
        for p in paths:
            _check_cancel(cancel_event)
            merger.append(os.path.abspath(p))
        merger.write(output_file)
        logger.info("Zapisano PDF: %s", output_file)
        return os.path.abspath(output_file)
    except Exception as exc:
        raise GenerationError(f"Błąd scalania PDF: {exc}") from exc
    finally:
        merger.close()


def _require_reportlab():
    if canvas is None:
        raise GenerationError(
            "Brak biblioteki reportlab. Zainstaluj ją w środowisku (pip install reportlab), aby dodać stopkę."
        )


def _get_footer_font_name() -> str:
    global _FOOTER_FONT_IN_USE
    _require_reportlab()
    if _FOOTER_FONT_IN_USE:
        return _FOOTER_FONT_IN_USE

    font_name = FOOTER_FONT
    for path in FOOTER_PREFERRED_FONT_PATHS:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(font_name, path))
                _FOOTER_FONT_IN_USE = font_name
                return _FOOTER_FONT_IN_USE
            except Exception:
                continue

    _FOOTER_FONT_IN_USE = FOOTER_FALLBACK_FONT
    return _FOOTER_FONT_IN_USE


def _create_footer_overlay(page_width: float, page_height: float, page_number: int) -> BytesIO:
    font_name = _get_footer_font_name()
    overlay = BytesIO()
    c = canvas.Canvas(overlay, pagesize=(page_width, page_height))
    c.setFont(font_name, FOOTER_FONT_SIZE)
    baseline = FOOTER_MARGIN_Y
    c.drawString(FOOTER_MARGIN_X + 10, baseline, FOOTER_LEFT_TEXT)
    c.drawCentredString(page_width / 2, baseline, str(page_number))
    c.drawRightString(page_width - FOOTER_MARGIN_X - 10, baseline, FOOTER_RIGHT_TEXT)
    c.save()
    overlay.seek(0)
    return overlay


def _apply_footer_to_pdf(pdf_path: str, cancel_event: Optional[threading.Event] = None):
    _require_reportlab()
    _check_cancel(cancel_event)
    if not os.path.exists(pdf_path):
        raise GenerationError(f"Nie znaleziono pliku PDF do oznaczenia stopką: {pdf_path}")

    reader = PdfReader(pdf_path)
    total_pages = len(reader.pages)
    if total_pages <= FOOTER_START_PAGE_INDEX + FOOTER_SKIP_LAST:
        return  # zbyt mały dokument, nie ma czego oznaczać

    writer = PdfWriter()
    for idx, page in enumerate(reader.pages):
        _check_cancel(cancel_event)
        page_number = idx + 1
        if idx >= FOOTER_START_PAGE_INDEX and idx < total_pages - FOOTER_SKIP_LAST:
            overlay_stream = _create_footer_overlay(float(page.mediabox.width), float(page.mediabox.height), page_number)
            overlay_pdf = PdfReader(overlay_stream)
            page.merge_page(overlay_pdf.pages[0])
        writer.add_page(page)

    temp_fd, temp_path = tempfile.mkstemp(prefix="wano-footer-", suffix=".pdf")
    try:
        with os.fdopen(temp_fd, "wb") as temp_file:
            writer.write(temp_file)
        shutil.move(temp_path, pdf_path)
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


def _find_soffice_binary() -> str:
    candidates = [
        os.environ.get("SOFFICE_PATH"),
        shutil.which("libreoffice"),
        shutil.which("soffice"),
        "/usr/bin/libreoffice",
        "/usr/bin/soffice",
        "/usr/local/bin/libreoffice",
        "/usr/local/bin/soffice",
    ]
    for candidate in candidates:
        if candidate and os.path.exists(candidate):
            return candidate
    raise GenerationError(
        "Brak libreoffice/soffice w PATH ani w standardowych lokalizacjach. "
        "Ustaw zmienną SOFFICE_PATH lub dodaj LibreOffice do PATH."
    )


def _convert_excel_to_pdf(
    excel_path: str, dest_dir: str, cancel_event: Optional[threading.Event] = None
) -> str:
    _check_cancel(cancel_event)
    os.makedirs(dest_dir, exist_ok=True)
    excel_abs = os.path.abspath(excel_path)
    out_name = os.path.splitext(os.path.basename(excel_abs))[0] + ".pdf"
    out_path = os.path.join(dest_dir, out_name)

    soffice = _find_soffice_binary()

    cmd = [
        soffice,
        "--headless",
        "--convert-to",
        "pdf",
        "--outdir",
        dest_dir,
        excel_abs,
    ]
    env = os.environ.copy()
    default_paths = [
        "/usr/local/sbin",
        "/usr/local/bin",
        "/usr/sbin",
        "/usr/bin",
        "/sbin",
        "/bin",
    ]
    env_paths = env.get("PATH", "").split(":")
    for p in default_paths:
        if p not in env_paths:
            env_paths.append(p)
    env["PATH"] = ":".join(env_paths)
    env["HOME"] = env.get("HOME", "/tmp")

    proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, env=env)
    stdout = b""
    stderr = b""
    try:
        while True:
            try:
                stdout, stderr = proc.communicate(timeout=0.5)
                break
            except subprocess.TimeoutExpired:
                _check_cancel(cancel_event)
                continue
    except GenerationCancelled:
        proc.terminate()
        try:
            proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            proc.kill()
        raise

    if proc.returncode != 0:
        stderr_msg = stderr.decode(errors="ignore") if stderr else stdout.decode(errors="ignore")
        raise GenerationError(f"Konwersja przez libreoffice nie powiodła się: {stderr_msg}")

    _check_cancel(cancel_event)
    if not os.path.exists(out_path):
        raise GenerationError(f"Nie znaleziono wyjściowego PDF: {out_path}")
    return out_path


def _export_sheet_uno(
    excel_path: str, sheet_name: str, target_pdf: str, cancel_event: Optional[threading.Event] = None
) -> Optional[str]:
    """Eksport arkusza przez UNO, usuwając inne arkusze i zachowując grafiki."""
    _check_cancel(cancel_event)
    try:
        import uno
        import unohelper
        from com.sun.star.beans import PropertyValue
    except ImportError:
        logger.warning("UNO not available; skipping UNO export for %s", sheet_name)
        return None

    soffice = _find_soffice_binary()
    _check_cancel(cancel_event)
    os.makedirs(os.path.dirname(os.path.abspath(target_pdf)) or ".", exist_ok=True)
    tmp_profile = tempfile.mkdtemp(prefix="lo-profile-")
    office_cmd = [
        soffice,
        "--headless",
        "--nologo",
        "--nodefault",
        "--nofirststartwizard",
        f"-env:UserInstallation=file://{tmp_profile}",
        "--accept=socket,host=127.0.0.1,port=2002;urp;",
    ]
    office_proc = subprocess.Popen(office_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

    def _connect():
        local_ctx = uno.getComponentContext()
        resolver = local_ctx.ServiceManager.createInstanceWithContext("com.sun.star.bridge.UnoUrlResolver", local_ctx)
        return resolver.resolve("uno:socket,host=127.0.0.1,port=2002;urp;StarOffice.ComponentContext")

    ctx = None
    for _ in range(40):
        try:
            _check_cancel(cancel_event)
            ctx = _connect()
            break
        except Exception:
            import time

            time.sleep(0.1)
    if ctx is None:
        office_proc.terminate()
        return None

    smgr = ctx.getServiceManager()
    desktop = smgr.createInstanceWithContext("com.sun.star.frame.Desktop", ctx)

    source_url = unohelper.systemPathToFileUrl(os.path.abspath(excel_path))
    load_props = (PropertyValue("Hidden", 0, True, 0),)
    _check_cancel(cancel_event)
    doc = desktop.loadComponentFromURL(source_url, "_blank", 0, load_props)
    if doc is None:
        desktop.terminate()
        office_proc.terminate()
        return None

    single_path = None
    try:
        _check_cancel(cancel_event)
        sheets = doc.Sheets
        if not sheets.hasByName(sheet_name):
            return None
        # Usuń wszystkie arkusze poza docelowym
        for name in list(sheets.ElementNames):
            _check_cancel(cancel_event)
            if name != sheet_name:
                try:
                    sheets.removeByName(name)
                except Exception:
                    pass
        # Ustaw aktywny arkusz na jedyny
        doc.CurrentController.setActiveSheet(sheets.getByName(sheet_name))

        # Zapisz kopię z jednym arkuszem
        temp_dir = tempfile.mkdtemp(prefix="wano-uno-")
        single_path = os.path.join(temp_dir, os.path.basename(excel_path))
        single_url = unohelper.systemPathToFileUrl(single_path)
        save_props = (PropertyValue("FilterName", 0, "Calc Office Open XML", 0),)
        doc.storeToURL(single_url, save_props)
    except Exception as exc:
        logger.warning("UNO export failed (prep) for %s: %s", sheet_name, exc)
        try:
            doc.close(True)
        except Exception:
            pass
        try:
            desktop.terminate()
        except Exception:
            pass
        office_proc.terminate()
        return None
    finally:
        try:
            doc.close(True)
        except Exception:
            pass
        try:
            desktop.terminate()
        except Exception:
            pass
        office_proc.terminate()

    # Konwersja tej kopii do PDF (cały workbook ma już jeden arkusz)
    if single_path:
        try:
            pdf_generated = _convert_excel_to_pdf(single_path, os.path.dirname(target_pdf), cancel_event)
            if os.path.exists(pdf_generated):
                shutil.move(pdf_generated, target_pdf)
                return target_pdf
        except Exception as exc:
            logger.warning("UNO copy convert failed for %s: %s", sheet_name, exc)
    return None



def _export_sheets(
    language_token: str,
    excel_path: Optional[str] = None,
    progress_cb=None,
    cancel_event: Optional[threading.Event] = None,
    register_cleanup: Optional[Callable[[str], None]] = None,
) -> List[str]:
    if load_workbook is None:
        raise GenerationError("Brak biblioteki openpyxl. Zainstaluj ją w venv (`pip install openpyxl`).")

    os.makedirs(EXPORT_DIR, exist_ok=True)
    source_excel = os.path.abspath(excel_path or EXCEL_PATH)

    try:
        wb = load_workbook(source_excel, keep_vba=True)
    except Exception as exc:
        raise GenerationError(f"Nie można otworzyć skoroszytu: {exc}") from exc

    matched_sheets: List[tuple[str, str]] = []
    sheet_patterns = re.compile(rf"^(\d+{language_token})")
    for sheet in wb.sheetnames:
        name = sheet.strip()
        match = sheet_patterns.match(name)
        if match:
            matched_sheets.append((match.group(1), sheet))

    if not matched_sheets:
        raise GenerationError(f"Nie znaleziono arkuszy pasujących do wzorca *{language_token} w skoroszycie.")

    matched_sheets.sort(key=lambda item: int(re.match(r"(\d+)", item[0]).group(1)))

    total = len(matched_sheets)
    with tempfile.TemporaryDirectory() as temp_dir:
        exported_prefixes: List[str] = []
        for idx, (prefix, sheet_name) in enumerate(matched_sheets, start=1):
            _check_cancel(cancel_event)
            layout_pdf = os.path.join(PDFY_DIR, f"{prefix}.pdf")
            num_prefix = int(re.match(r"(\d+)", prefix).group(1))
            layout_required = num_prefix != 1
            if layout_required and not os.path.exists(layout_pdf):
                logger.warning("Pomijam arkusz %s - brak pliku układu: %s", sheet_name, layout_pdf)
                if progress_cb:
                    progress_cb("export", min(15 + int((idx / total) * 70), 90), f"Pomijam {sheet_name}")
                continue
            if not layout_required and not os.path.exists(layout_pdf):
                logger.info("Używam Start%s jako layout dla %s", language_token, sheet_name)

            temp_pdf_path = os.path.join(temp_dir, f"{prefix}ex.pdf")
            pdf_path = _export_sheet_uno(
                source_excel,
                sheet_name,
                temp_pdf_path,
                cancel_event=cancel_event,
            )
            if pdf_path is None:
                _check_cancel(cancel_event)
                temp_wb = load_workbook(source_excel, keep_vba=True)
                if sheet_name not in temp_wb.sheetnames:
                    continue
                for other in list(temp_wb.sheetnames):
                    _check_cancel(cancel_event)
                    if other != sheet_name:
                        temp_wb.remove(temp_wb[other])
                temp_wb.active = temp_wb[sheet_name]
                temp_excel_path = os.path.join(temp_dir, f"{prefix}.xlsm")
                temp_wb.save(temp_excel_path)
                pdf_path = _convert_excel_to_pdf(temp_excel_path, temp_dir, cancel_event)

            final_pdf = os.path.join(EXPORT_DIR, f"{prefix}ex.pdf")
            shutil.move(pdf_path, final_pdf)
            if register_cleanup:
                register_cleanup(final_pdf)

            exported_prefixes.append(prefix)
            if progress_cb:
                progress_cb("export", min(15 + int((idx / total) * 70), 90), f"Wyeksportowano {sheet_name}")

    if not exported_prefixes:
        raise GenerationError(
            f"Brak arkuszy z plikiem układu w {PDFY_DIR}. Upewnij się, że pliki *.pdf istnieją (np. 2{language_token}.pdf)."
        )

    return exported_prefixes


def generate_price_list(
    language: str,
    excel_path: Optional[str] = None,
    progress_cb=None,
    cancel_event: Optional[threading.Event] = None,
    register_cleanup: Optional[Callable[[str], None]] = None,
) -> str:
    """Generuje cennik PDF (Linux, libreoffice).

    progress_cb(stage, percent, message) — opcjonalny callback do raportowania postępu.
    """
    language = language.lower()
    if language not in {"pl", "en"}:
        raise GenerationError("Język musi być 'pl' lub 'en'.")

    token = "PL" if language == "pl" else "EN"
    output_file = OUTPUT_FILE_PL if language == "pl" else OUTPUT_FILE_EN

    _check_cancel(cancel_event)
    if progress_cb:
        progress_cb("start", 5, "Start generowania")

    _validate_assets(token, excel_path)
    _check_cancel(cancel_event)

    if progress_cb:
        progress_cb("export", 10, "Eksport arkuszy")
    prefixes = _export_sheets(
        token,
        excel_path,
        progress_cb,
        cancel_event=cancel_event,
        register_cleanup=register_cleanup,
    )

    _check_cancel(cancel_event)
    if progress_cb:
        progress_cb("merge", 92, "Scalanie PDF")
    parts = [os.path.join(PDFY_DIR, f"Start{token}.pdf")]
    for prefix in prefixes:
        _check_cancel(cancel_event)
        sheet_pdf = os.path.join(EXPORT_DIR, f"{prefix}ex.pdf")
        layout_pdf = os.path.join(PDFY_DIR, f"{prefix}.pdf")
        if not os.path.exists(sheet_pdf):
            raise GenerationError(f"Brak wygenerowanego PDF: {sheet_pdf}")
        # Kolejność: Start -> 1ex -> 2.pdf + 2ex -> ... -> End
        num_prefix = int(re.match(r"(\d+)", prefix).group(1))
        if num_prefix == 1:
            parts.append(sheet_pdf)
            continue
        if not os.path.exists(layout_pdf):
            raise GenerationError(f"Brak pliku układu: {layout_pdf}")
        parts.append(layout_pdf)
        parts.append(sheet_pdf)
    parts.append(os.path.join(PDFY_DIR, f"End{token}.pdf"))

    target_dir = OUTPUT_DIR_PL if language == "pl" else OUTPUT_DIR_EN
    os.makedirs(target_dir, exist_ok=True)
    result = _merge_pdf_list(parts, output_file, cancel_event=cancel_event)
    if progress_cb:
        progress_cb("merge", 96, "Dodawanie stopki")
    _apply_footer_to_pdf(result, cancel_event=cancel_event)
    if register_cleanup:
        register_cleanup(result)
    if progress_cb:
        progress_cb("done", 100, "Gotowe")
    return result


if __name__ == "__main__":
    import argparse

    logging.basicConfig(level=logging.INFO)
    parser = argparse.ArgumentParser(description="Generator cennika WANO (wersja web/CLI, libreoffice).")
    parser.add_argument("--lang", choices=["pl", "en"], default="pl", help="Wybierz język PDF.")
    parser.add_argument("--excel", default=None, help="Ścieżka do pliku Excela.")
    args = parser.parse_args()

    try:
        output = generate_price_list(args.lang, args.excel)
        print(f"✅ Wygenerowano: {output}")
    except GenerationError as exc:
        print(f"❌ {exc}")
